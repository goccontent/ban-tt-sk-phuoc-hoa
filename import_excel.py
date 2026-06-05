"""
Đọc file Excel kế hoạch sự kiện.
Cột bắt buộc: Ngày + Sự kiện/Chương trình (tự nhận diện tên cột).
Cột tuỳ chọn: Giờ (giờ bắt đầu sự kiện) — dùng làm deadline nhịp "Trong".
"""
import re
from datetime import date, datetime, time as _time

from openpyxl import load_workbook

DATE_HEADERS = ("ngày", "ngay", "date", "thời gian", "thoi gian")
NAME_HEADERS = (
    "sự kiện", "su kien", "chương trình", "chuong trinh",
    "lễ", "le", "tên sự kiện", "ten su kien", "nội dung", "noi dung",
    "sự kiện / chương trình", "su kien / chuong trinh",
)
# "Giờ" là cột tuỳ chọn; chỉ nhận diện qua từ khoá giờ đặc trưng để không
# nhầm với cột "Thời gian" (vốn được hiểu là cột Ngày).
TIME_HEADERS = ("giờ", "gio", "time", "giờ bắt đầu", "gio bat dau")


def _norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def _find_columns(header_row):
    date_col = name_col = time_col = None
    for idx, cell in enumerate(header_row):
        h = _norm(cell)
        if not h:
            continue
        if time_col is None and any(k in h for k in TIME_HEADERS):
            time_col = idx
            continue  # cột Giờ không kiêm cột Ngày/Sự kiện
        if date_col is None and any(k in h for k in DATE_HEADERS):
            date_col = idx
        if name_col is None and any(k in h for k in NAME_HEADERS):
            name_col = idx
    return date_col, name_col, time_col


def _parse_time_cell(value):
    """Trả về 'HH:MM' hoặc None. Chấp nhận ô giờ Excel, '10:00', '10h30', '8g', '7 giờ'."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, _time):
        return value.strftime("%H:%M")
    s = str(value).strip().lower()
    # 10:00 · 10.00 · 10h30 · 10g30 · 10 giờ 30 · 8h · 7g · 9 giờ
    m = re.match(r"^(\d{1,2})\s*(?:[:.h g]|giờ|gio)\s*(\d{1,2})?", s)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2)) if m.group(2) else 0
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"
    m = re.match(r"^(\d{1,2})$", s)
    if m:
        hh = int(m.group(1))
        if 0 <= hh <= 23:
            return f"{hh:02d}:00"
    return None


def _parse_date_cell(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    s = str(value).strip()
    # 15.03.2026 hoặc 15/03/2026
    m = re.match(r"(\d{1,2})[./](\d{1,2})[./](\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{int(d):02d}.{int(mo):02d}.{y}"
    # 03.06–03.07.2026 hoặc 03.06-03.07.2026
    m = re.match(
        r"(\d{1,2})[./](\d{1,2})\s*[-–]\s*(\d{1,2})[./](\d{1,2})[./](\d{4})",
        s,
    )
    if m:
        d1, m1, d2, m2, y = m.groups()
        return f"{int(d1):02d}.{int(m1):02d}–{int(d2):02d}.{int(m2):02d}.{y}"
    return s


def parse_excel(file_path_or_stream):
    wb = load_workbook(file_path_or_stream, read_only=True, data_only=True)
    imported = []
    errors = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = None
        date_col = name_col = time_col = None
        for i, row in enumerate(rows[:10]):
            dc, nc, tc = _find_columns(row)
            if dc is not None and nc is not None:
                header_idx = i
                date_col, name_col, time_col = dc, nc, tc
                break

        if header_idx is None:
            errors.append(f"Sheet '{sheet.title}': không tìm thấy cột Ngày và Sự kiện")
            continue

        for row in rows[header_idx + 1 :]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            name = row[name_col] if name_col < len(row) else None
            raw_date = row[date_col] if date_col < len(row) else None
            name = str(name).strip() if name else ""
            date_str = _parse_date_cell(raw_date)
            if not name:
                continue
            if not date_str:
                errors.append(f"Bỏ qua '{name}': thiếu ngày")
                continue
            rec = {
                "name": name,
                "date": date_str,
                "sheet": sheet.title,
            }
            if time_col is not None:
                raw_time = row[time_col] if time_col < len(row) else None
                t = _parse_time_cell(raw_time)
                if t:
                    rec["time"] = t
            imported.append(rec)

    wb.close()
    return imported, errors


def to_event_records(parsed_rows, existing_ids=None):
    from deadline_rules import parse_event_date_range  # noqa: PLC0415
    from events_store import make_event_id  # noqa: PLC0415

    existing_ids = set(existing_ids or [])
    events = []
    for row in parsed_rows:
        eid = make_event_id(row["name"], row["date"], existing_ids | {e["id"] for e in events})
        rng = parse_event_date_range(row["date"])
        rec = {
            "id": eid,
            "name": row["name"],
            "date": row["date"],
        }
        if rng:
            start, end = rng
            # eventDate = ngày kết thúc (như data.js đọc endISO); eventStartDate = ngày bắt đầu
            rec["eventDate"] = end.isoformat()
            rec["eventStartDate"] = start.isoformat()
        if row.get("time"):
            # giờ bắt đầu sự kiện — neo deadline nhịp "Trong"
            rec["eventStartTime"] = row["time"]
        events.append(rec)
    return events
