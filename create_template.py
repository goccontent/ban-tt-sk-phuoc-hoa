"""Tạo file mẫu Excel kế hoạch sự kiện."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT = Path(__file__).parent / "mau-ke-hoach-su-kien.xlsx"

SAMPLE = [
    ("STT", "Ngày", "Sự kiện / Chương trình", "Ghi chú"),
    (1, "06.06.2026", "Khóa Bồi dưỡng Truyền thông Công giáo", ""),
    (2, "25.06.2026", "Thông báo chuẩn bị Rước Lễ Lần Đầu", ""),
    (3, "28.06.2026", "Thánh Lễ Lãnh nhận Bí Tích Rước Lễ Lần Đầu", ""),
    (4, "03.06–03.07.2026", "Chương trình Hè 2026", "Cả tháng"),
]

wb = Workbook()
ws = wb.active
ws.title = "Kế hoạch"

header_fill = PatternFill("solid", fgColor="1A4D6D")
header_font = Font(bold=True, color="FFFFFF")

for r, row in enumerate(SAMPLE, 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if r == 1:
            cell.fill = header_fill
            cell.font = header_font

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 20

wb.save(OUT)
print("Created mau-ke-hoach-su-kien.xlsx")
